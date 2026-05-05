from __future__ import annotations

import argparse
import math
import re
import struct
from collections import OrderedDict
from pathlib import Path
from typing import Dict, Iterable, List

import pandas as pd


NULL_SENTINEL = -1.0e30
PAGE_SIZE = 2048
OUT_COPY_DATA_BLOCK_SIZE = 4096
FIELD_DESC_SIZE = 28
HEADER_OFFSET = 0x70
DEFAULT_FLOAT_PRECISION = 6

# Non-Datamine files that should still appear in the batch summary.
# .dm files are parsed using the Datamine schema reader below.
# .csv/.xlsx files are treated as tabular mirrors/exports: their headers are
# read so the classifier can apply the same column-signature rules as for .dm.
SUMMARY_REFERENCE_SUFFIXES = {
    ".mac",      # Datamine macro files
    ".htm",      # Datamine recorded HTML script files
    ".html",     # Optional alias, retained for robustness
    ".jpg",
    ".jpeg",     # Optional alias for JPG images
    ".png",
    ".bmp",
    ".pdf",
}
SUMMARY_TABLE_SUFFIXES = {
    ".csv",
    ".xlsx",
}
SUMMARY_SUPPORTED_SUFFIXES = {".dm"}.union(SUMMARY_REFERENCE_SUFFIXES).union(SUMMARY_TABLE_SUFFIXES)

SUMMARY_KIND_BY_SUFFIX = {
    ".dm": "Datamine_DM_File",
    ".csv": "Tabular_Export_CSV_File",
    ".xlsx": "Tabular_Export_XLSX_File",
    ".mac": "Datamine_Macro_File",
    ".htm": "Datamine_HTML_Script_File",
    ".html": "Datamine_HTML_Script_File",
    ".jpg": "Image_or_PDF_File",
    ".jpeg": "Image_or_PDF_File",
    ".png": "Image_or_PDF_File",
    ".bmp": "Image_or_PDF_File",
    ".pdf": "Image_or_PDF_File",
}

# Only extension-only file groups get a fixed destination at summary stage.
# CSV/XLSX exports are deliberately left for the classifier because they can
# carry the same columns as their source .dm files.
SUMMARY_SUGGESTED_FOLDER_BY_SUFFIX = {
    ".mac": "34_Macros",
    ".htm": "35_Scripts",
    ".html": "35_Scripts",
    ".jpg": "36_Images",
    ".jpeg": "36_Images",
    ".png": "36_Images",
    ".bmp": "36_Images",
    ".pdf": "36_Images",
}

CSV_ENCODINGS_TO_TRY = ("utf-8-sig", "utf-8", "cp1252", "latin1")

# openpyxl/Excel rejects most ASCII control characters in text cells.
# Datamine alphanumeric fields can contain these when binary bytes are decoded.
ILLEGAL_XLSX_CHAR_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def _f32(bs: bytes) -> float:
    return struct.unpack("<f", bs)[0]

def _f64(bs: bytes) -> float:
    return struct.unpack("<d", bs)[0]


def parse_legacy_dm_schema(dm_file: Path) -> dict:
    raw = dm_file.read_bytes()
    if len(raw) < PAGE_SIZE:
        raise ValueError(f"{dm_file} is too small to be a valid legacy Datamine .dm file")

    def _parse_compact_header() -> dict:
        n_fields = int(_f32(raw[0x64:0x68]))
        file_counter = int(_f32(raw[0x68:0x6C]))
        record_len_header = int(_f32(raw[0x6C:0x70]))

        if not (0 < n_fields <= 4096):
            raise ValueError(f"compact header not valid: n_fields={n_fields}")

        desc = []
        off = HEADER_OFFSET
        for _ in range(n_fields):
            chunk = raw[off:off + FIELD_DESC_SIZE]
            if len(chunk) < FIELD_DESC_SIZE:
                raise ValueError(f"Header ended unexpectedly in {dm_file}")

            name = chunk[0:8].decode("ascii", errors="ignore").rstrip()
            field_type = chunk[8:12].decode("ascii", errors="ignore").rstrip()
            meta = struct.unpack("<4f", chunk[12:28])

            desc.append({
                "name": name,
                "type": field_type,
                "meta": meta,
                "layout": "out_copy",
            })
            off += FIELD_DESC_SIZE

        data_start = math.ceil(off / PAGE_SIZE) * PAGE_SIZE
        row_descriptors = [d for d in desc if d["meta"][0] > 0]

        row_words = len(row_descriptors)
        row_bytes = row_words * 4
        rows_per_page = PAGE_SIZE // row_bytes if row_bytes else 0
        page_slack = PAGE_SIZE - (rows_per_page * row_bytes) if row_bytes else PAGE_SIZE
        physical_page_count = len(raw) // PAGE_SIZE
        physical_data_page_count = (len(raw) - data_start) // PAGE_SIZE
        header_page_count = data_start // PAGE_SIZE

        # In compact/single-precision Datamine files, the value at 0x68 normally
        # stores the allocated page count recorded by Datamine. Some legacy files
        # are physically longer than this value because stale pages remain at the
        # end of the file. Use the header page count where it is plausible.
        if (
            file_counter is not None
            and file_counter > header_page_count
            and file_counter <= physical_page_count
        ):
            page_count = file_counter - header_page_count
            trailing_allocated_pages_ignored = physical_data_page_count - page_count
        else:
            page_count = physical_data_page_count
            trailing_allocated_pages_ignored = 0

        page_count = max(0, min(page_count, physical_data_page_count))

        # For this compact/single-precision Datamine layout, the value at 0x6C
        # is not the row byte length. It is the number of active records stored
        # in the final used data page. The physical file is page-allocated, so
        # unused slots in the last page may contain stale bytes from previous data.
        if page_count > 0 and 0 < record_len_header <= rows_per_page:
            last_page_active_rows = record_len_header
            active_row_count = ((page_count - 1) * rows_per_page) + last_page_active_rows
        else:
            last_page_active_rows = None
            active_row_count = None

        return {
            "n_fields_header": n_fields,
            "file_counter_header": file_counter,
            "record_len_header": record_len_header,
            "last_page_active_rows": last_page_active_rows,
            "active_row_count": active_row_count,
            "all_descriptors": desc,
            "row_descriptors": row_descriptors,
            "data_start": data_start,
            "row_words": row_words,
            "row_bytes": row_bytes,
            "rows_per_page": rows_per_page,
            "page_slack": page_slack,
            "page_count": page_count,
            "physical_data_page_count": physical_data_page_count,
            "trailing_allocated_pages_ignored": trailing_allocated_pages_ignored,
            "file_size": len(raw),
            "header_layout": "compact",
            "data_block_size": PAGE_SIZE,
        }

    def _parse_out_copy_header() -> dict:
        # Observed on files beginning with text like:
        # "OUT ... FILE CREATED USING COPY ..."
        # Count and descriptor metadata are stored as doubles.
        n_fields = int(struct.unpack("<d", raw[0xC8:0xD0])[0])
        record_len_header = int(struct.unpack("<d", raw[0xD0:0xD8])[0])
        

        if not (0 < n_fields <= 4096):
            raise ValueError(f"out/copy header not valid: n_fields={n_fields}")

        desc = []
        off = 0xE0
        field_desc_size = 56

        for _ in range(n_fields):
            chunk = raw[off:off + field_desc_size]
            if len(chunk) < field_desc_size:
                raise ValueError(f"Header ended unexpectedly in {dm_file}")

            # Name occupies the first 16 bytes, often with internal spaces between chunks.
            # Remove spaces to recover fields such as ZONET, DESCRIP, XMORIG, DENSITY, etc.
            name = chunk[0:16].decode("ascii", errors="ignore").replace("\x00", "")
            name = "".join(name.split())   # remove embedded fixed-width spaces

            field_type = chunk[16:20].decode("ascii", errors="ignore").strip()

            field_len = struct.unpack("<I", chunk[20:24])[0]

            meta = struct.unpack("<4d", chunk[24:56])

            desc.append({
                "name": name,
                "type": field_type,
                "field_len": field_len,
                "meta": meta,
                "layout": "out_copy",
            })
            off += field_desc_size

        # OUT/COPY files observed here start data on a 4096-byte boundary
        # and pack explicit records into 4096-byte data blocks.
        data_start = math.ceil(off / OUT_COPY_DATA_BLOCK_SIZE) * OUT_COPY_DATA_BLOCK_SIZE

        # For this header family, meta[0] > 0 identifies explicit stored row fields.
        # meta[0] == 0 is used for implicit/system fields such as XMORIG, NX, etc.
        row_descriptors = [d for d in desc if d["meta"][0] > 0]
        row_words = len(row_descriptors)
        row_bytes = row_words * 8
        data_block_size = OUT_COPY_DATA_BLOCK_SIZE
        rows_per_page = data_block_size // row_bytes if row_bytes else 0
        page_slack = data_block_size - (rows_per_page * row_bytes) if row_bytes else data_block_size
        page_count = (len(raw) - data_start) // data_block_size

        return {
            "n_fields_header": n_fields,
            "file_counter_header": None,
            "record_len_header": record_len_header,
            "all_descriptors": desc,
            "row_descriptors": row_descriptors,
            "data_start": data_start,
            "row_words": row_words,
            "row_bytes": row_bytes,
            "rows_per_page": rows_per_page,
            "page_slack": page_slack,
            "page_count": page_count,
            "file_size": len(raw),
            "header_layout": "out_copy",
            "row_layout_known": True,
            "data_block_size": data_block_size,
        }

    # Try the original compact layout first.
    try:
        return _parse_compact_header()
    except Exception:
        pass

    # Fallback for the OUT/COPY layout.
    return _parse_out_copy_header()

def norm_field_type(ftype: str) -> str:
    if ftype is None:
        return ""
    t = "".join(ch for ch in str(ftype).upper() if ch.isalpha())
    return t[:1]

def logical_columns(descriptors: List[Dict]) -> List[str]:
    cols = []
    i = 0
    while i < len(descriptors):
        d = descriptors[i]
        name = d["name"]
        ftype = norm_field_type(d.get("type", ""))

        cols.append(d["name"])

        if ftype == "N":
            i += 1
        elif ftype == "A":
            j = i + 1
            while (
                j < len(descriptors)
                and norm_field_type(descriptors[j].get("type", "")) == "A"
                and descriptors[j]["name"] == name
            ):
                j += 1
            i = j
        else:
            i += 1
    return cols


def _looks_like_split_4byte_alpha_word(word: bytes) -> bool:
    """
    Some OUT/COPY-style files store numeric fields as 8-byte doubles, but
    alphanumeric chunks still use only the first 4 bytes of each 8-byte word.

    In those files the fifth byte is commonly NUL and the remaining bytes may
    contain duplicate/garbage characters. Example: b"UNMI\x00ED " should be
    decoded as "UNMI", not "UNMIED".
    """
    if len(word) != 8:
        return False

    first_half_has_text = any(b not in (0, 32) for b in word[:4])
    return first_half_has_text and word[4] == 0


def _detect_alpha_bytes_per_word(
    raw: bytes,
    data_start: int,
    data_block_size: int,
    page_count: int,
    rows_per_page: int,
    row_bytes: int,
    row_descriptors: List[Dict],
    word_size: int,
    sample_rows: int = 200,
) -> int:
    """Return how many bytes of each fixed-width alphanumeric word hold text."""
    if word_size != 8:
        return word_size

    alpha_positions = [
        i for i, d in enumerate(row_descriptors)
        if norm_field_type(d.get("type", "")) == "A"
    ]
    if not alpha_positions:
        return word_size

    checked = 0
    split_hits = 0

    for p in range(page_count):
        page_start = data_start + p * data_block_size
        page = raw[page_start : page_start + data_block_size]

        for r in range(rows_per_page):
            row_start = r * row_bytes
            row = page[row_start : row_start + row_bytes]
            if len(row) != row_bytes:
                continue

            # Ignore entirely blank rows when detecting layout.
            if not row.strip(b"\x00 "):
                continue

            checked += 1

            for pos in alpha_positions:
                s = pos * word_size
                word = row[s : s + word_size]
                if _looks_like_split_4byte_alpha_word(word):
                    split_hits += 1

            if checked >= sample_rows:
                return 4 if split_hits > 0 else word_size

    return 4 if split_hits > 0 else word_size


def _decode_row(
    chunk: bytes,
    row_descriptors: List[Dict],
    word_size: int = 4,
    alpha_bytes_per_word: int | None = None,
) -> OrderedDict:
    out = OrderedDict()
    pos = 0

    if alpha_bytes_per_word is None:
        alpha_bytes_per_word = word_size

    while pos < len(row_descriptors):
        d = row_descriptors[pos]
        name = d["name"]
        ftype = norm_field_type(d.get("type", ""))

        start = pos * word_size
        end = (pos + 1) * word_size

        if ftype == "N":
            if word_size == 4:
                val = _f32(chunk[start:end])
            elif word_size == 8:
                val = _f64(chunk[start:end])
            else:
                raise ValueError(f"Unsupported numeric word size: {word_size}")

            if abs(val - NULL_SENTINEL) < 1e25:
                val = None

            out[name] = val
            pos += 1

        elif ftype == "A":
            parts = []
            while (
                pos < len(row_descriptors)
                and norm_field_type(row_descriptors[pos].get("type", "")) == "A"
                and row_descriptors[pos]["name"] == name
            ):
                s = pos * word_size
                e = s + alpha_bytes_per_word

                # Each alphanumeric word is fixed-width padded. Strip trailing
                # padding from each word before joining, otherwise fields such
                # as UNMINED become "UNMI    NED". For mixed 8-byte numeric /
                # 4-byte alpha files, only the first four bytes are decoded.
                part = chunk[s:e].decode("ascii", errors="ignore").replace("\x00", "")
                part = ILLEGAL_XLSX_CHAR_RE.sub("", part)
                parts.append(part.rstrip())
                pos += 1

            txt = "".join(parts).strip()
            out[name] = txt

        else:
            raise ValueError(f"Unsupported normalized field type: {d.get('type')}")

    return out


def _is_blank_explicit_dataframe_row(df: pd.DataFrame) -> pd.Series:
    """Return True for rows where all explicit stored fields are blank/zero."""
    if df.empty:
        return pd.Series(dtype=bool)

    obj_cols = df.select_dtypes(include="object").columns.tolist()
    num_cols = [c for c in df.columns if c not in obj_cols]

    blank_num = (
        df[num_cols].fillna(0).eq(0).all(axis=1)
        if num_cols
        else pd.Series(True, index=df.index)
    )
    blank_obj = (
        df[obj_cols].fillna("").apply(lambda s: s.eq("").all(), axis=1)
        if obj_cols
        else pd.Series(True, index=df.index)
    )
    return blank_num & blank_obj


def _implicit_value_from_descriptor(d: Dict):
    """Return the constant value for an implicit descriptor."""
    ftype = norm_field_type(d.get("type", ""))

    if ftype == "N":
        val = d.get("meta", (None, None, None, None))[3]
        if val is not None and abs(val - NULL_SENTINEL) < 1e25:
            return None
        return val

    if ftype == "A":
        return ""

    return None


def _expand_with_implicit_fields(df_explicit: pd.DataFrame, all_descriptors: List[Dict]) -> pd.DataFrame:
    """
    Rebuild the dataframe in Datamine logical field order.

    OUT/COPY files store only explicit fields in each row. Implicit/system
    fields, e.g. XMORIG, NX, NY and NZ, are stored in the header as constants
    and must be reinserted into the exported table.
    """
    if df_explicit.empty:
        return pd.DataFrame(columns=logical_columns(all_descriptors))

    expanded_rows = []

    for row in df_explicit.to_dict(orient="records"):
        out = OrderedDict()
        i = 0

        while i < len(all_descriptors):
            d = all_descriptors[i]
            name = d["name"]
            ftype = norm_field_type(d.get("type", ""))

            if ftype == "N":
                if d["meta"][0] > 0:
                    out[name] = row.get(name)
                else:
                    out[name] = _implicit_value_from_descriptor(d)
                i += 1

            elif ftype == "A":
                j = i + 1
                group = [d]
                while (
                    j < len(all_descriptors)
                    and norm_field_type(all_descriptors[j].get("type", "")) == "A"
                    and all_descriptors[j]["name"] == name
                ):
                    group.append(all_descriptors[j])
                    j += 1

                if any(g["meta"][0] > 0 for g in group):
                    out[name] = row.get(name, "")
                else:
                    out[name] = _implicit_value_from_descriptor(d)

                i = j

            else:
                i += 1

        expanded_rows.append(out)

    return pd.DataFrame(expanded_rows)


def clean_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Remove control characters that openpyxl cannot write to XLSX."""
    out = df.copy()
    obj_cols = out.select_dtypes(include=["object"]).columns

    for col in obj_cols:
        out[col] = out[col].map(
            lambda v: ILLEGAL_XLSX_CHAR_RE.sub("", v) if isinstance(v, str) else v
        )

    return out


def read_legacy_dm(dm_file: Path, drop_blank_rows: bool = True) -> pd.DataFrame:
    schema = parse_legacy_dm_schema(dm_file)
    raw = dm_file.read_bytes()

    row_desc = schema["row_descriptors"]
    data_start = schema["data_start"]
    page_count = schema["page_count"]
    header_layout = schema.get("header_layout", "compact")

    if header_layout == "compact":
        word_size = 4
        data_block_size = PAGE_SIZE
    elif header_layout == "out_copy":
        word_size = 8
        data_block_size = schema.get("data_block_size", OUT_COPY_DATA_BLOCK_SIZE)
    else:
        raise ValueError(f"Unsupported header layout: {header_layout}")

    row_words = len(row_desc)
    row_bytes = schema.get("row_bytes")
    if row_bytes is None:
        row_bytes = row_words * word_size

    rows_per_page = schema.get("rows_per_page")
    if rows_per_page is None:
        rows_per_page = data_block_size // row_bytes if row_bytes else 0

    if page_count is None:
        page_count = (len(raw) - data_start) // data_block_size

    alpha_bytes_per_word = _detect_alpha_bytes_per_word(
        raw=raw,
        data_start=data_start,
        data_block_size=data_block_size,
        page_count=page_count,
        rows_per_page=rows_per_page,
        row_bytes=row_bytes,
        row_descriptors=row_desc,
        word_size=word_size,
    )

    active_row_count = schema.get("active_row_count") if header_layout == "compact" else None

    rows = []
    rows_read = 0
    for p in range(page_count):
        page = raw[data_start + p * data_block_size : data_start + (p + 1) * data_block_size]

        for r in range(rows_per_page):
            if active_row_count is not None and rows_read >= active_row_count:
                break

            chunk = page[r * row_bytes : (r + 1) * row_bytes]
            if len(chunk) != row_bytes:
                continue
            rows.append(
                _decode_row(
                    chunk,
                    row_desc,
                    word_size=word_size,
                    alpha_bytes_per_word=alpha_bytes_per_word,
                )
            )
            rows_read += 1

        if active_row_count is not None and rows_read >= active_row_count:
            break

    df = pd.DataFrame(rows)

    if drop_blank_rows and not df.empty:
        df = df.loc[~_is_blank_explicit_dataframe_row(df)].reset_index(drop=True)

    if header_layout == "out_copy":
        df = _expand_with_implicit_fields(df, schema["all_descriptors"])

    return df

def round_numeric_columns(df: pd.DataFrame, decimals: int = DEFAULT_FLOAT_PRECISION) -> pd.DataFrame:
    out = df.copy()
    num_cols = out.select_dtypes(include=["number"]).columns
    if len(num_cols) > 0:
        out[num_cols] = out[num_cols].round(decimals)
    return out


def export_csv(
    dm_file: Path,
    csv_file: Path,
    drop_blank_rows: bool = True,
    decimals: int = DEFAULT_FLOAT_PRECISION,
) -> Path:
    df = read_legacy_dm(dm_file, drop_blank_rows=drop_blank_rows)
    df = round_numeric_columns(df, decimals=decimals)
    csv_file.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(csv_file, index=False, float_format=f"%.{decimals}f")
    return csv_file


def export_xlsx(
    dm_file: Path,
    xlsx_file: Path,
    drop_blank_rows: bool = True,
    decimals: int = DEFAULT_FLOAT_PRECISION,
    sheet_name: str = "data",
) -> Path:
    df = read_legacy_dm(dm_file, drop_blank_rows=drop_blank_rows)
    df = round_numeric_columns(df, decimals=decimals)
    df = clean_dataframe_for_excel(df)
    xlsx_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(xlsx_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        # Apply a numeric format to numeric cells so Excel opens them as numbers
        num_cols = df.select_dtypes(include=["number"]).columns.tolist()
        num_col_idx = {name: i + 1 for i, name in enumerate(df.columns)}
        number_format = f"0.{''.join(['0'] * decimals)}" if decimals > 0 else "0"

        for col_name in num_cols:
            col_idx = num_col_idx[col_name]
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=col_idx).number_format = number_format

    return xlsx_file


def print_summary(dm_file: Path) -> None:
    schema = parse_legacy_dm_schema(dm_file)
    cols = logical_columns(schema["all_descriptors"])
    print(f"File: {dm_file}")
    print(f"File size: {schema['file_size']} bytes")
    print(f"Header descriptors: {schema['n_fields_header']}")
    print(f"Header counter value: {schema['file_counter_header']}")
    print(f"Header record_len value: {schema['record_len_header']}")
    print(f"Row words: {schema['row_words']}")
    print(f"Row bytes: {schema['row_bytes']}")
    print(f"Data block size: {schema.get('data_block_size', PAGE_SIZE)} bytes")
    print(f"Rows per data block: {schema['rows_per_page']}")
    print(f"Block slack bytes: {schema['page_slack']}")
    print(f"Data blocks: {schema['page_count']}")
    print("Logical columns:")
    for i, c in enumerate(cols, start=1):
        print(f"  {i}. {c}")


def iter_dm_files(path: Path, recursive: bool = False) -> Iterable[Path]:
    """Yield Datamine .dm files only. Used by export routines."""
    if path.is_file():
        if path.suffix.lower() == ".dm":
            yield path
        return

    pattern = "**/*.dm" if recursive else "*.dm"
    for p in path.glob(pattern):
        if p.is_file():
            yield p


def iter_summary_files(path: Path, recursive: bool = False) -> Iterable[Path]:
    """Yield all file types that should appear in the batch summary."""
    if path.is_file():
        if path.suffix.lower() in SUMMARY_SUPPORTED_SUFFIXES:
            yield path
        return

    pattern = "**/*" if recursive else "*"
    for p in path.glob(pattern):
        if p.is_file() and p.suffix.lower() in SUMMARY_SUPPORTED_SUFFIXES:
            yield p



def _safe_stem_name(dm_file: Path) -> str:
    return dm_file.stem


def _normalise_column_names(columns: Iterable[object]) -> List[str]:
    """Return clean string column names for summary/classifier use."""
    out: List[str] = []
    for col in columns:
        text = str(col).strip()
        if text and not text.lower().startswith("unnamed:"):
            out.append(text)
    return out


def _count_csv_data_rows(csv_file: Path) -> int | None:
    """Estimate CSV data rows as physical lines minus header.

    This is deliberately cheap and avoids loading large Datamine exports into
    memory. It may over-count only where a CSV has embedded multiline text.
    Datamine table exports normally do not use multiline text fields.
    """
    try:
        with csv_file.open("rb") as f:
            line_count = sum(1 for _ in f)
        return max(0, line_count - 1)
    except Exception:
        return None


def _read_csv_table_summary(csv_file: Path) -> Dict[str, object]:
    """Read only the CSV header and row-count estimate."""
    errors = []

    for encoding in CSV_ENCODINGS_TO_TRY:
        try:
            # sep=None with the Python engine lets pandas infer comma/tab/etc.
            # Datamine CSV exports are usually comma-delimited, but this makes
            # the summary more tolerant of manually saved files.
            header_df = pd.read_csv(
                csv_file,
                nrows=0,
                encoding=encoding,
                sep=None,
                engine="python",
            )
            columns = _normalise_column_names(header_df.columns)
            return {
                "columns_list": columns,
                "tabular_rows": _count_csv_data_rows(csv_file),
                "tabular_columns": len(columns),
                "csv_encoding": encoding,
            }
        except Exception as exc:
            errors.append(f"{encoding}: {exc}")

    raise ValueError("Could not read CSV header; " + " | ".join(errors))


def _read_xlsx_table_summary(xlsx_file: Path) -> Dict[str, object]:
    """Read workbook sheet names, selected sheet header and rough dimensions."""
    excel = pd.ExcelFile(xlsx_file)
    sheet_names = excel.sheet_names
    if not sheet_names:
        raise ValueError("Workbook contains no sheets")

    # Datamine exports created by this script use 'data'. Otherwise use the
    # first worksheet, which is the safest default for simple workbook exports.
    sheet_name = "data" if "data" in sheet_names else sheet_names[0]
    header_df = pd.read_excel(xlsx_file, sheet_name=sheet_name, nrows=0)
    columns = _normalise_column_names(header_df.columns)

    tabular_rows = None
    tabular_columns = len(columns)

    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_file, read_only=True, data_only=True)
        ws = wb[sheet_name]
        if ws.max_row is not None:
            tabular_rows = max(0, int(ws.max_row) - 1)
        if ws.max_column is not None:
            tabular_columns = int(ws.max_column)
        wb.close()
    except Exception:
        # Header extraction is the important part for classification. Dimension
        # detection is useful but not critical.
        pass

    return {
        "columns_list": columns,
        "tabular_rows": tabular_rows,
        "tabular_columns": tabular_columns,
        "sheet_name": sheet_name,
        "workbook_sheets": "; ".join(sheet_names),
    }


def _build_dm_source_indexes(files: Iterable[Path]) -> tuple[Dict[tuple[str, str], List[Path]], Dict[str, List[Path]]]:
    """Build lookup indexes for same-stem .dm source matching."""
    by_parent_and_stem: Dict[tuple[str, str], List[Path]] = {}
    by_stem: Dict[str, List[Path]] = {}

    for file_path in files:
        if file_path.suffix.lower() != ".dm":
            continue
        stem_key = file_path.stem.lower()
        parent_key = str(file_path.parent.resolve()).lower()
        by_parent_and_stem.setdefault((parent_key, stem_key), []).append(file_path)
        by_stem.setdefault(stem_key, []).append(file_path)

    return by_parent_and_stem, by_stem


def _source_match_fields(
    file_path: Path,
    by_parent_and_stem: Dict[tuple[str, str], List[Path]],
    by_stem: Dict[str, List[Path]],
) -> Dict[str, object]:
    """Return same-stem source .dm matching fields for CSV/XLSX mirror files."""
    suffix = file_path.suffix.lower()
    stem_key = file_path.stem.lower()

    base = {
        "mirror_source_stem": file_path.stem,
        "mirror_of_dm": False,
        "source_dm_file": "",
        "source_match_status": "not_applicable",
    }

    if suffix not in SUMMARY_TABLE_SUFFIXES:
        return base

    parent_key = str(file_path.parent.resolve()).lower()
    same_dir_matches = by_parent_and_stem.get((parent_key, stem_key), [])
    if len(same_dir_matches) == 1:
        base.update({
            "mirror_of_dm": True,
            "source_dm_file": str(same_dir_matches[0]),
            "source_match_status": "same_stem_dm_found_same_folder",
        })
        return base
    if len(same_dir_matches) > 1:
        base.update({
            "mirror_of_dm": True,
            "source_dm_file": "; ".join(str(p) for p in same_dir_matches),
            "source_match_status": "multiple_same_stem_dm_found_same_folder",
        })
        return base

    any_folder_matches = by_stem.get(stem_key, [])
    if len(any_folder_matches) == 1:
        base.update({
            "mirror_of_dm": True,
            "source_dm_file": str(any_folder_matches[0]),
            "source_match_status": "same_stem_dm_found_elsewhere_in_scan",
        })
        return base
    if len(any_folder_matches) > 1:
        base.update({
            "mirror_of_dm": True,
            "source_dm_file": "; ".join(str(p) for p in any_folder_matches),
            "source_match_status": "multiple_same_stem_dm_found_in_scan",
        })
        return base

    base["source_match_status"] = "same_stem_dm_not_found"
    return base


def _base_summary_row(
    file_path: Path,
    status: str,
    by_parent_and_stem: Dict[tuple[str, str], List[Path]] | None = None,
    by_stem: Dict[str, List[Path]] | None = None,
) -> Dict[str, object]:
    """Return a summary row with the same schema used for parsed .dm files."""
    suffix = file_path.suffix.lower()
    if by_parent_and_stem is None:
        by_parent_and_stem = {}
    if by_stem is None:
        by_stem = {}

    row = {
        "file": str(file_path),
        "file_name": file_path.name,
        "extension": suffix,
        "file_size_bytes": file_path.stat().st_size if file_path.exists() else None,
        "summary_file_kind": SUMMARY_KIND_BY_SUFFIX.get(suffix),
        "summary_suggested_folder": SUMMARY_SUGGESTED_FOLDER_BY_SUFFIX.get(suffix),
        "header_descriptors": None,
        "header_layout": None,
        "row_layout_known": None,
        "row_words": None,
        "row_bytes": None,
        "rows_per_page": None,
        "data_pages": None,
        "active_rows_from_header": None,
        "last_page_active_rows": None,
        "logical_columns": None,
        "columns": "",
        "tabular_rows": None,
        "tabular_columns": None,
        "sheet_name": "",
        "workbook_sheets": "",
        "csv_encoding": "",
        "status": status,
    }
    row.update(_source_match_fields(file_path, by_parent_and_stem, by_stem))
    return row


def _summarise_dm_file(dm_file: Path) -> Dict[str, object]:
    schema = parse_legacy_dm_schema(dm_file)
    cols = logical_columns(schema["all_descriptors"])
    suffix = dm_file.suffix.lower()
    row = _base_summary_row(dm_file, status="ok")
    row.update({
        "summary_file_kind": SUMMARY_KIND_BY_SUFFIX.get(suffix),
        "summary_suggested_folder": None,
        "file_size_bytes": schema["file_size"],
        "header_descriptors": schema["n_fields_header"],
        "header_layout": schema.get("header_layout"),
        "row_layout_known": schema.get("row_layout_known"),
        "row_words": schema["row_words"],
        "row_bytes": schema["row_bytes"],
        "rows_per_page": schema["rows_per_page"],
        "data_pages": schema["page_count"],
        "active_rows_from_header": schema.get("active_row_count"),
        "last_page_active_rows": schema.get("last_page_active_rows"),
        "logical_columns": len(cols),
        "columns": ", ".join(cols),
    })
    return row


def _summarise_table_file(
    file_path: Path,
    by_parent_and_stem: Dict[tuple[str, str], List[Path]],
    by_stem: Dict[str, List[Path]],
) -> Dict[str, object]:
    suffix = file_path.suffix.lower()
    row = _base_summary_row(file_path, status="ok", by_parent_and_stem=by_parent_and_stem, by_stem=by_stem)

    if suffix == ".csv":
        info = _read_csv_table_summary(file_path)
    elif suffix == ".xlsx":
        info = _read_xlsx_table_summary(file_path)
    else:
        raise ValueError(f"Unsupported table extension: {suffix}")

    cols = info.pop("columns_list")
    row.update(info)
    row.update({
        "logical_columns": len(cols),
        "columns": ", ".join(cols),
    })
    return row


def _summarise_reference_file(
    file_path: Path,
    by_parent_and_stem: Dict[tuple[str, str], List[Path]],
    by_stem: Dict[str, List[Path]],
) -> Dict[str, object]:
    suffix = file_path.suffix.lower()
    if suffix not in SUMMARY_REFERENCE_SUFFIXES:
        return _base_summary_row(
            file_path,
            status=f"SKIPPED: unsupported extension {suffix}",
            by_parent_and_stem=by_parent_and_stem,
            by_stem=by_stem,
        )
    return _base_summary_row(file_path, status="ok", by_parent_and_stem=by_parent_and_stem, by_stem=by_stem)


def batch_summary(path: Path, recursive: bool = False) -> pd.DataFrame:
    summary_files = list(iter_summary_files(path, recursive=recursive))
    by_parent_and_stem, by_stem = _build_dm_source_indexes(summary_files)

    rows = []
    for file_path in summary_files:
        try:
            suffix = file_path.suffix.lower()
            if suffix == ".dm":
                rows.append(_summarise_dm_file(file_path))
            elif suffix in SUMMARY_TABLE_SUFFIXES:
                rows.append(_summarise_table_file(file_path, by_parent_and_stem, by_stem))
            else:
                rows.append(_summarise_reference_file(file_path, by_parent_and_stem, by_stem))
        except Exception as exc:
            rows.append(
                _base_summary_row(
                    file_path,
                    status=f"ERROR: {exc}",
                    by_parent_and_stem=by_parent_and_stem,
                    by_stem=by_stem,
                )
            )

    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(by=["extension", "file"], kind="stable").reset_index(drop=True)
    return out

def batch_export(
    path: Path,
    output_dir: Path,
    recursive: bool = False,
    decimals: int = DEFAULT_FLOAT_PRECISION,
    drop_blank_rows: bool = True,
    export_csv_flag: bool = True,
    export_xlsx_flag: bool = True,
) -> pd.DataFrame:
    rows = []
    output_dir.mkdir(parents=True, exist_ok=True)

    for dm_file in iter_dm_files(path, recursive=recursive):
        base_name = _safe_stem_name(dm_file)
        csv_path = output_dir / f"{base_name}.csv"
        xlsx_path = output_dir / f"{base_name}.xlsx"

        row = {
            "file": str(dm_file),
            "csv_file": str(csv_path) if export_csv_flag else None,
            "xlsx_file": str(xlsx_path) if export_xlsx_flag else None,
            "rows_exported": None,
            "columns_exported": None,
            "status": None,
        }

        try:
            df = read_legacy_dm(dm_file, drop_blank_rows=drop_blank_rows)
            df = round_numeric_columns(df, decimals=decimals)

            if export_csv_flag:
                df.to_csv(csv_path, index=False, float_format=f"%.{decimals}f")

            if export_xlsx_flag:
                df_xlsx = clean_dataframe_for_excel(df)
                with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                    df_xlsx.to_excel(writer, index=False, sheet_name="data")
                    ws = writer.book["data"]
                    num_cols = df_xlsx.select_dtypes(include=["number"]).columns.tolist()
                    num_col_idx = {name: i + 1 for i, name in enumerate(df_xlsx.columns)}
                    number_format = f"0.{''.join(['0'] * decimals)}" if decimals > 0 else "0"

                    for col_name in num_cols:
                        col_idx = num_col_idx[col_name]
                        for row_idx in range(2, len(df) + 2):
                            ws.cell(row=row_idx, column=col_idx).number_format = number_format

            row["rows_exported"] = len(df)
            row["columns_exported"] = len(df.columns)
            row["status"] = "ok"

        except Exception as exc:
            row["status"] = f"ERROR: {exc}"

        rows.append(row)

    return pd.DataFrame(rows)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Read legacy Datamine .dm files directly, including chunked alphanumeric fields, with batch CSV/XLSX export and mixed-file batch summaries."
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_summary = sub.add_parser("summary", help="Print schema summary for one .dm file")
    p_summary.add_argument("dm_file", type=Path)

    p_export = sub.add_parser("export", help="Export one .dm file")
    p_export.add_argument("dm_file", type=Path)
    p_export.add_argument("output_base", type=Path, help="Output path without extension, or with .csv/.xlsx")
    p_export.add_argument("--csv-only", action="store_true")
    p_export.add_argument("--xlsx-only", action="store_true")
    p_export.add_argument("--keep-blank-rows", action="store_true")
    p_export.add_argument("--decimals", type=int, default=DEFAULT_FLOAT_PRECISION)

    p_batch_summary = sub.add_parser("batch-summary", help="Summarise .dm, .csv, .xlsx, .mac, .htm/html, image and PDF files in a folder")
    p_batch_summary.add_argument("path", type=Path)
    p_batch_summary.add_argument("--recursive", action="store_true")
    p_batch_summary.add_argument("--out", type=Path, default=None)

    p_batch_export = sub.add_parser("batch-export", help="Export all .dm files in a folder")
    p_batch_export.add_argument("path", type=Path)
    p_batch_export.add_argument("output_dir", type=Path)
    p_batch_export.add_argument("--recursive", action="store_true")
    p_batch_export.add_argument("--csv-only", action="store_true")
    p_batch_export.add_argument("--xlsx-only", action="store_true")
    p_batch_export.add_argument("--keep-blank-rows", action="store_true")
    p_batch_export.add_argument("--decimals", type=int, default=DEFAULT_FLOAT_PRECISION)
    p_batch_export.add_argument("--log", type=Path, default=None)

    args = parser.parse_args()

    if args.cmd == "summary":
        print_summary(args.dm_file)

    elif args.cmd == "export":
        export_csv_flag = not args.xlsx_only
        export_xlsx_flag = not args.csv_only

        output_base = args.output_base
        if output_base.suffix.lower() == ".csv":
            csv_path = output_base
            xlsx_path = output_base.with_suffix(".xlsx")
        elif output_base.suffix.lower() == ".xlsx":
            xlsx_path = output_base
            csv_path = output_base.with_suffix(".csv")
        else:
            csv_path = output_base.with_suffix(".csv")
            xlsx_path = output_base.with_suffix(".xlsx")

        if export_csv_flag:
            export_csv(
                args.dm_file,
                csv_path,
                drop_blank_rows=not args.keep_blank_rows,
                decimals=args.decimals,
            )
            print(f"Exported CSV: {csv_path}")

        if export_xlsx_flag:
            export_xlsx(
                args.dm_file,
                xlsx_path,
                drop_blank_rows=not args.keep_blank_rows,
                decimals=args.decimals,
            )
            print(f"Exported XLSX: {xlsx_path}")

    elif args.cmd == "batch-summary":
        df = batch_summary(args.path, recursive=args.recursive)
        if args.out:
            args.out.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(args.out, index=False)
            print(f"Summary written to: {args.out}")
        else:
            print(df.to_string(index=False))

    elif args.cmd == "batch-export":
        export_csv_flag = not args.xlsx_only
        export_xlsx_flag = not args.csv_only

        df = batch_export(
            args.path,
            args.output_dir,
            recursive=args.recursive,
            decimals=args.decimals,
            drop_blank_rows=not args.keep_blank_rows,
            export_csv_flag=export_csv_flag,
            export_xlsx_flag=export_xlsx_flag,
        )

        if args.log:
            args.log.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(args.log, index=False)
            print(f"Log written to: {args.log}")
        else:
            print(df.to_string(index=False))


if __name__ == "__main__":
    main()
